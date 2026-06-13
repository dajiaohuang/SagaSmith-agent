from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


CELL_REF = re.compile(r"(?:(?:'([^']+)'|([\w\u4e00-\u9fff]+))!)?\$?([A-Z]{1,3})\$?(\d+)")
CORE_SHEETS = {"主要情况", "BUY点计算器"}


def extract_formula_catalog(template: Path) -> dict:
    workbook = load_workbook(template, data_only=False)
    formulas = []
    sheet_stats = {}
    for worksheet in workbook.worksheets:
        cross_dependencies: Counter[str] = Counter()
        sheet_formulas = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                formula = str(cell.value)
                dependencies = []
                for quoted, plain, column, row_number in CELL_REF.findall(formula):
                    sheet = quoted or plain or worksheet.title
                    reference = f"{sheet}!{column}{row_number}"
                    dependencies.append(reference)
                    if sheet != worksheet.title:
                        cross_dependencies[sheet] += 1
                category = "core_rule_candidate" if worksheet.title in CORE_SHEETS else "lookup_or_template_formula"
                item = {
                    "sheet": worksheet.title,
                    "cell": cell.coordinate,
                    "formula": formula,
                    "dependencies": dependencies,
                    "category": category,
                }
                formulas.append(item)
                sheet_formulas.append(item)
        sheet_stats[worksheet.title] = {
            "formula_count": len(sheet_formulas),
            "cross_sheet_dependencies": dict(cross_dependencies.most_common()),
        }
    return {
        "template": template.name,
        "formula_count": len(formulas),
        "sheet_stats": sheet_stats,
        "formulas": formulas,
    }


def write_report(catalog: dict, target: Path) -> None:
    lines = [
        "# DND 5E 人物卡模板公式审计",
        "",
        f"- 模板：`{catalog['template']}`",
        f"- 公式总数：`{catalog['formula_count']}`",
        "",
        "## 工作表统计",
        "",
        "| 工作表 | 公式数 | 跨表依赖 |",
        "|---|---:|---|",
    ]
    for sheet, stats in catalog["sheet_stats"].items():
        dependencies = ", ".join(f"{key}: {value}" for key, value in stats["cross_sheet_dependencies"].items()) or "-"
        lines.append(f"| {sheet} | {stats['formula_count']} | {dependencies} |")
    lines.extend([
        "",
        "## 代码化策略",
        "",
        "- 直接代码化：属性调整值、熟练加值、职业豁免、技能、先攻、AC、HP、被动察觉、法术 DC/命中、负重、27 点购点。",
        "- 规则目录保留：职业/种族特性查表、装备查表、法术位和法术列表联动。",
        "- 保留在 Excel：纯展示、单元格镜像和大型法术表引用。",
        "",
        "## 核心公式候选",
        "",
    ])
    for item in catalog["formulas"]:
        if item["category"] == "core_rule_candidate":
            lines.append(f"- `{item['sheet']}!{item['cell']}`: `{item['formula']}`")
    target.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("template", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    catalog = extract_formula_catalog(args.template)
    (args.output / "formula_catalog.json").write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_report(catalog, args.output / "formula_report.md")
    print(f"Extracted {catalog['formula_count']} formulas to {args.output}")


if __name__ == "__main__":
    main()
