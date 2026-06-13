from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CLASS_NAMES = [
    "bard", "cleric", "druid", "paladin", "ranger", "sorcerer",
    "warlock", "wizard", "artificer", "chronurgy_wizard", "graviturgy_wizard",
]
SPELL_FIELDS = [
    "level", "school", "ritual", "casting_time", "range", "verbal", "somatic",
    "material", "material_description", "duration", "description",
]


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u3000", " ").split()).strip()


def spell_key(chinese_name: str, english_name: str) -> str:
    return (english_name or chinese_name).casefold().replace(" ", "")


def parse_spell_row(row: tuple, source: str, priority: int, layout: str) -> dict | None:
    if layout == "template":
        chinese_name, english_name = normalize_text(row[1]), normalize_text(row[14])
        values = row[2:13]
        classes = row[15:26]
        publication = ""
    else:
        english_name, chinese_name = normalize_text(row[0]), normalize_text(row[1])
        values = row[2:13]
        classes = row[13:24]
        publication = normalize_text(row[24]) if len(row) > 24 else ""
    if not chinese_name and not english_name:
        return None
    spell = {
        "id": spell_key(chinese_name, english_name),
        "name": chinese_name,
        "english_name": english_name,
        "classes": [name for name, marker in zip(CLASS_NAMES, classes, strict=False) if normalize_text(marker)],
        "publication": publication,
        "sources": [source],
        "_priority": priority,
    }
    spell.update({field: normalize_text(value) for field, value in zip(SPELL_FIELDS, values, strict=True)})
    return spell


def merge_spell(current: dict | None, incoming: dict) -> dict:
    if not current:
        return incoming
    winner, other = (incoming, current) if incoming["_priority"] > current["_priority"] else (current, incoming)
    merged = dict(winner)
    for field in ["name", "english_name", *SPELL_FIELDS, "publication"]:
        if not merged.get(field):
            merged[field] = other.get(field, "")
    merged["classes"] = sorted(set(current.get("classes", [])) | set(incoming.get("classes", [])))
    merged["sources"] = list(dict.fromkeys([*current.get("sources", []), *incoming.get("sources", [])]))
    merged["_priority"] = max(current["_priority"], incoming["_priority"])
    return merged


def build_spell_catalog(raw_dir: Path) -> list[dict]:
    merged: dict[str, dict] = {}
    character_template = next(raw_dir.glob("*人物卡模板.xlsx"), None)
    if character_template:
        workbook = load_workbook(character_template, read_only=True, data_only=True)
        for row in workbook.worksheets[4].iter_rows(min_row=2, values_only=True):
            spell = parse_spell_row(row, f"{character_template.name}:法术大全", 20, "template")
            if spell:
                merged[spell["id"]] = merge_spell(merged.get(spell["id"]), spell)

    standalone = next(raw_dir.glob("*法术大全*.xlsx"), None)
    if standalone:
        workbook = load_workbook(standalone, read_only=True, data_only=True)
        priorities = [40, 10, 5]
        for worksheet, priority in zip(workbook.worksheets[:3], priorities, strict=True):
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                spell = parse_spell_row(row, f"{standalone.name}:{worksheet.title.strip()}", priority, "standalone")
                if spell:
                    merged[spell["id"]] = merge_spell(merged.get(spell["id"]), spell)
    result = []
    for spell in merged.values():
        spell.pop("_priority", None)
        result.append(spell)
    return sorted(result, key=lambda item: (level_number(item.get("level")), item.get("name") or item.get("english_name")))


def level_number(value: Any) -> int:
    match = re.search(r"\d+", str(value or ""))
    return int(match.group()) if match else 0


def save_spell_catalog(raw_dir: Path, target: Path) -> list[dict]:
    spells = build_spell_catalog(raw_dir)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(spells, ensure_ascii=False, indent=2), encoding="utf-8")
    return spells


@lru_cache(maxsize=4)
def load_spell_catalog(data_dir: str) -> list[dict]:
    root = Path(data_dir)
    target = root / "generated" / "spells" / "spell_catalog.json"
    if target.exists():
        return json.loads(target.read_text(encoding="utf-8"))
    return save_spell_catalog(root / "raw", target)


def search_spells(query: str, data_dir: Path, limit: int = 5, class_name: str | None = None) -> list[dict]:
    query = normalize_text(query).casefold()
    if not query:
        return []
    terms = [term for term in re.findall(r"[\w\u4e00-\u9fff]+", query) if len(term) > 1]
    ranked = []
    for spell in load_spell_catalog(str(data_dir.resolve())):
        if class_name and class_name.casefold() not in spell.get("classes", []):
            continue
        chinese = spell.get("name", "").casefold()
        english = spell.get("english_name", "").casefold()
        names = f"{chinese} {english}"
        searchable = " ".join([
            names, spell.get("school", ""), spell.get("description", ""),
            spell.get("material_description", ""), " ".join(spell.get("classes", [])),
        ]).casefold()
        score = 0
        if query in {chinese, english}:
            score += 100
        elif query in names:
            score += 50
        for term in terms:
            if term in names:
                score += 20
            elif term in searchable:
                score += 2
        if score:
            ranked.append((score, spell))
    ranked.sort(key=lambda item: (-item[0], level_number(item[1].get("level")), item[1].get("name", "")))
    return [{**spell, "score": score} for score, spell in ranked[:limit]]


def enrich_character_spells(spells: list[Any], data_dir: Path) -> list[dict]:
    enriched = []
    for item in spells:
        name = item if isinstance(item, str) else item.get("name") or item.get("english_name")
        matches = search_spells(str(name or ""), data_dir, 1)
        if matches:
            enriched.append(matches[0])
        elif isinstance(item, dict):
            enriched.append(item)
        else:
            enriched.append({"name": str(item), "unresolved": True})
    return enriched


def format_spell(spell: dict) -> str:
    components = "".join(key for key, field in [("V", "verbal"), ("S", "somatic"), ("M", "material")] if spell.get(field))
    classes = "、".join(spell.get("classes", [])) or "未标注"
    return (
        f"{spell.get('name') or spell.get('english_name')} ({spell.get('english_name')})\n"
        f"{spell.get('level')}环 {spell.get('school')} | {spell.get('casting_time')} | {spell.get('range')}\n"
        f"成分：{components or '-'} | 持续：{spell.get('duration')} | 职业：{classes}\n"
        f"{spell.get('description')}"
    ).strip()


def extract_direct_spell_query(message: str) -> str | None:
    text = normalize_text(message)
    prefixes = ("/法术 ", "/spell ", "查询法术 ", "查法术 ", "法术查询 ")
    for prefix in prefixes:
        if text.casefold().startswith(prefix.casefold()):
            return text[len(prefix):].strip() or None
    return None


def direct_spell_lookup(message: str, data_dir: Path, limit: int = 5) -> tuple[str, list[dict]] | None:
    explicit = extract_direct_spell_query(message)
    if explicit:
        return explicit, search_spells(explicit, data_dir, limit)
    text = normalize_text(message)
    if not any(marker in text.casefold() for marker in ["是什么", "效果", "介绍", "怎么用", "查询", "what is"]):
        return None
    folded = text.casefold()
    exact = [
        spell for spell in load_spell_catalog(str(data_dir.resolve()))
        if (spell.get("name") and spell["name"] in text)
        or (spell.get("english_name") and spell["english_name"].casefold() in folded)
    ]
    exact.sort(key=lambda spell: max(len(spell.get("name", "")), len(spell.get("english_name", ""))), reverse=True)
    return (text, exact[:limit]) if exact else None
